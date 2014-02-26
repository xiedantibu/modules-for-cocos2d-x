from openpyxl import load_workbook
from biplist import *
import sys
import json

def work(excel_file_name):
	wb = load_workbook(excel_file_name)
	sheetnames = wb.get_sheet_names()
	ret=[]
	for name in sheetnames:
		ws = wb.get_sheet_by_name(name)
		pair={}
		for r in range(ws.get_highest_row()):
			key = ws.cell(row = r, column = 0).value
			value = ws.cell(row = r, column = 1).value
			pair[key] = value
		ret.append((ws.title,pair))
	return ret

if __name__=='__main__':
	if(len(sys.argv)<3):
		print('Usage: python excel_to_dict.py [excel_file_name] [dest_file_type:plist|json]')
	else:
		data = work(sys.argv[1])
		if(sys.argv[2].lower()=='json'):
			for (file_name, dic) in data:
				f = open(file_name+'.json','w')
				c = json.dumps(dic, ensure_ascii=False, indent=2)
				if(isinstance(c, unicode)):
					 f.write(c.encode('utf8'))
				else:
					f.write(c)
				f.close()
		elif(sys.argv[2].lower()=='plist'):
			for (file_name, dic) in data:
				try:
					writePlist(dic, file_name+'.plist')
				except (InvalidPlistException, NotBinaryPlistException), e:
					print(e)
		elif(sys.argv[2].lower()=='sqlite'):
			print('Warning: excel of dictionary cannot convert to sqlite')
		else:
			print("Error: wrong dest_file_type")